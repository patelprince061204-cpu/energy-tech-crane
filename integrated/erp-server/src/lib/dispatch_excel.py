#!/usr/bin/env python3
"""
Generates real .xlsx dispatch forms using openpyxl.
Called from Node.js via child_process.spawn with JSON on stdin.
Writes the .xlsx bytes to stdout so Node reads them directly.
"""
import sys
import io
import json
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Colour palette matching the ERP theme ──────────────────────────────────
NAVY   = "1C2530"   # header background
WHITE  = "FFFFFF"
AMBER  = "F59E0B"   # section accent
LIGHT  = "F1F5F9"   # alternating row
BORDER = "CBD5E1"

def thin_border(positions=("top","bottom","left","right")):
    sides = {p: Side(style="thin", color=BORDER) for p in positions}
    return Border(**{p: sides[p] for p in positions})

def all_border():
    s = Side(style="thin", color=BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill():
    return PatternFill("solid", fgColor=NAVY)

def amber_fill():
    return PatternFill("solid", fgColor=AMBER)

def light_fill():
    return PatternFill("solid", fgColor=LIGHT)

def bold_white(size=11):
    return Font(bold=True, color=WHITE, size=size)

def bold(size=11):
    return Font(bold=True, size=size)

def normal(size=10):
    return Font(size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left_mid():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def set_row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ── FORM 1: Dispatch Material List ─────────────────────────────────────────
MATERIAL_ITEMS = [
    "Main Girder", "Wire Rope Hoist", "Crab Unit Assembly", "End Carriage",
    "C-Rail", "C-Rail Accessories", "C-Rail Angle", "DSL Bus Bar",
    "DSL Accessories", "DSL Angle", "Rubber Buffer",
    "HT Motor with brake with gear box", "CT motor with brake with gear box",
    "LT motor with brake with gear box (1)", "LT motor with brake with gear box (2)",
    "Panel", "Bolt", "Limit Switch", "Wire Rope", "Hook",
    "D-Clamp", "Push Button", "Wireless Control Remote", "Stiker",
    "Square Bar", "Panel Angle", "LT murga channel", "CT murga channel",
    "Cable", "Touching Color", "CT & LT Limit switch",
    "Up & Down Limit switch (Rotary & Roller)",
]

def build_material_list_xlsx(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dispatch Material List"

    # Column widths
    set_col_width(ws, 1, 8)    # Sr No
    set_col_width(ws, 2, 48)   # Description
    set_col_width(ws, 3, 22)   # Quantity

    r = 1

    # ── Title ──
    ws.merge_cells(f"A{r}:C{r}")
    ws[f"A{r}"] = "DISPATCH MATERIAL LIST"
    ws[f"A{r}"].font = Font(bold=True, size=16, color=NAVY)
    ws[f"A{r}"].alignment = center()
    set_row_height(ws, r, 32)
    r += 1

    # ── Blank spacer ──
    set_row_height(ws, r, 6)
    r += 1

    # ── Header info block ──
    info_fields = [
        ("COMPANY NAME",   data.get("company_name", "")),
        ("TRANSPORT NAME", data.get("transporter_name", "")),
        ("VEHICLE NUMBER", data.get("vehicle_number", "")),
        ("DATE",           data.get("dispatch_date", "")),
        ("DISPATCH NUMBER",data.get("dispatch_number", "")),
        ("PO NUMBER",      data.get("so_number", "")),
    ]
    for label, value in info_fields:
        ws.merge_cells(f"A{r}:B{r}")
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=True, size=10, color=NAVY)
        ws[f"A{r}"].alignment = left_mid()
        ws[f"C{r}"] = value
        ws[f"C{r}"].font = Font(bold=True, size=11)
        ws[f"C{r}"].alignment = left_mid()
        ws[f"C{r}"].border = Border(bottom=Side(style="medium", color=NAVY))
        set_row_height(ws, r, 20)
        r += 1

    # ── Spacer ──
    set_row_height(ws, r, 8)
    r += 1

    # ── Table header ──
    headers = ["SR NO.", "DESCRIPTION", "QUANTITY"]
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=ci, value=h)
        cell.font = bold_white(11)
        cell.fill = header_fill()
        cell.alignment = center()
        cell.border = all_border()
    set_row_height(ws, r, 22)
    r += 1

    # ── Material rows ──
    for i, item in enumerate(MATERIAL_ITEMS):
        # Sr No
        c1 = ws.cell(row=r, column=1, value=i + 1)
        c1.font = normal(10)
        c1.alignment = center()
        c1.border = all_border()
        if i % 2 == 1:
            c1.fill = light_fill()

        # Description
        c2 = ws.cell(row=r, column=2, value=item)
        c2.font = normal(10)
        c2.alignment = left_mid()
        c2.border = all_border()
        if i % 2 == 1:
            c2.fill = light_fill()

        # Quantity (blank for user to fill)
        c3 = ws.cell(row=r, column=3, value="")
        c3.border = all_border()
        if i % 2 == 1:
            c3.fill = light_fill()

        set_row_height(ws, r, 18)
        r += 1

    # ── 5 extra blank rows ──
    for _ in range(5):
        for ci in range(1, 4):
            cell = ws.cell(row=r, column=ci, value="")
            cell.border = all_border()
            cell.fill = PatternFill("solid", fgColor="FEF9EE")  # faint amber tint
        set_row_height(ws, r, 18)
        r += 1

    # ── Spacer ──
    set_row_height(ws, r, 10)
    r += 1

    # ── Supervisor signature ──
    ws.merge_cells(f"A{r}:C{r}")
    ws[f"A{r}"] = "Supervisor Signature:  ___________________________________"
    ws[f"A{r}"].font = Font(bold=True, size=11, color=NAVY)
    ws[f"A{r}"].alignment = Alignment(horizontal="right", vertical="center")
    set_row_height(ws, r, 24)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── FORM 2: Hoist Material / L-Block / Girder ──────────────────────────────
HOIST_ROWS  = [
    "Main drum (HP/RPM/SR.NO.)", "Main drum (Motor Break)",
    "Hoist CT Motor", "Hoist CT Break", "Hoist CT Gear Box",
    "Drum Size", "Hoist Gear Box", "First Gear",
    "Second Gear", "Third Gear", "Fourth Gear",
]
LBLOCK_ROWS = [
    "L-Block wheel Assembly (OD & ID)",
    "BOX SIZE (TOP BOTTOM & WEB)", "COUPLING",
    "LT MOTOR", "LT GEAR BOX MODEL", "LT BREAK",
    "RUBBER BUFFER", "CONNECTION PLATE", "BOLT",
]
GIRDER_ROWS = [
    "GIRDER LENGTH (SPAN)", "TOP PLATE SIZE", "BOTTOM PLATE SIZE",
    "WEB PLATE SIZE", "FULL STIFFENER", "HALF STIFFENER", "ANGLE",
]

def write_section(ws, r, title, col_header, rows):
    """Write a coloured section header + column header + blank fillable rows."""
    # Section title spanning both columns
    ws.merge_cells(f"A{r}:B{r}")
    ws[f"A{r}"] = title
    ws[f"A{r}"].font = Font(bold=True, size=12, color=WHITE)
    ws[f"A{r}"].fill = amber_fill()
    ws[f"A{r}"].alignment = left_mid()
    ws[f"A{r}"].border = all_border()
    set_row_height(ws, r, 24)
    r += 1

    # Column sub-header
    ws[f"A{r}"] = "ITEM"
    ws[f"A{r}"].font = bold_white(10)
    ws[f"A{r}"].fill = header_fill()
    ws[f"A{r}"].alignment = center()
    ws[f"A{r}"].border = all_border()

    ws[f"B{r}"] = col_header
    ws[f"B{r}"].font = bold_white(10)
    ws[f"B{r}"].fill = header_fill()
    ws[f"B{r}"].alignment = center()
    ws[f"B{r}"].border = all_border()
    set_row_height(ws, r, 20)
    r += 1

    # Data rows
    for i, row_label in enumerate(rows):
        c1 = ws.cell(row=r, column=1, value=row_label)
        c1.font = Font(bold=True, size=10)
        c1.fill = light_fill() if i % 2 == 0 else PatternFill("solid", fgColor=WHITE)
        c1.alignment = left_mid()
        c1.border = all_border()

        c2 = ws.cell(row=r, column=2, value="")
        c2.fill = PatternFill("solid", fgColor=WHITE)
        c2.border = all_border()
        set_row_height(ws, r, 20)
        r += 1

    # Spacer
    set_row_height(ws, r, 8)
    r += 1
    return r


def build_hoist_material_xlsx(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hoist & Girder Details"

    set_col_width(ws, 1, 40)
    set_col_width(ws, 2, 36)

    r = 1

    # Title
    ws.merge_cells(f"A{r}:B{r}")
    ws[f"A{r}"] = "HOIST MATERIAL & GIRDER DETAILS"
    ws[f"A{r}"].font = Font(bold=True, size=15, color=NAVY)
    ws[f"A{r}"].alignment = center()
    set_row_height(ws, r, 30)
    r += 1

    set_row_height(ws, r, 6)
    r += 1

    # Dispatch info
    info = [
        ("Company",       data.get("company_name", "")),
        ("Dispatch #",    data.get("dispatch_number", "")),
        ("Date",          data.get("dispatch_date", "")),
        ("Vehicle",       data.get("vehicle_number", "")),
        ("Crane Type",    f"{data.get('crane_type','')} – {data.get('capacity','')}".strip(" –")),
    ]
    for label, value in info:
        ws[f"A{r}"] = label
        ws[f"A{r}"].font = Font(bold=True, size=10, color=NAVY)
        ws[f"A{r}"].alignment = left_mid()
        ws[f"B{r}"] = value
        ws[f"B{r}"].font = Font(bold=True, size=11)
        ws[f"B{r}"].alignment = left_mid()
        ws[f"B{r}"].border = Border(bottom=Side(style="medium", color=NAVY))
        set_row_height(ws, r, 20)
        r += 1

    set_row_height(ws, r, 10)
    r += 1

    r = write_section(ws, r, "Hoist Material",          "Specification / SR.NO.", HOIST_ROWS)
    r = write_section(ws, r, "L-Block End Carriage",    "Specification",          LBLOCK_ROWS)
    r = write_section(ws, r, "Main Box Girder Selection","Specification",         GIRDER_ROWS)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Entry point ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    raw = sys.stdin.buffer.read()
    data = json.loads(raw)
    form_type = data.get("form_type", "material_list")

    if form_type == "material_list":
        xlsx_bytes = build_material_list_xlsx(data)
    elif form_type == "hoist_material":
        xlsx_bytes = build_hoist_material_xlsx(data)
    else:
        sys.exit(1)

    sys.stdout.buffer.write(xlsx_bytes)
